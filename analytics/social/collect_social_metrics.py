#!/usr/bin/env python3
"""
Instagram + Threads social metrics collector.

This script is operational code, but it does not run automatically.
Run it manually only when metrics collection is needed.

Required environment variables:
- INSTAGRAM_ACCESS_TOKEN
- INSTAGRAM_IG_USER_ID
- THREADS_ACCESS_TOKEN
- THREADS_USER_ID

Optional:
- SOCIAL_METRICS_XLSX
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


GRAPH_BASE = "https://graph.facebook.com/v19.0"
THREADS_BASE = "https://graph.threads.net/v1.0"

INSTAGRAM_POST_METRICS = [
    "reach",
    "likes",
    "comments",
    "saved",
    "shares",
]

THREADS_POST_METRICS = [
    "views",
    "likes",
    "replies",
    "reposts",
    "quotes",
    "shares",
]


@dataclass
class ApiResult:
    ok: bool
    status: int
    data: dict[str, Any]


def mask(value: str | None) -> str:
    if not value:
        return ""
    if len(value) <= 8:
        return "*" * len(value)
    return f"{value[:4]}********{value[-4:]}"


def request_json(url: str, params: dict[str, str]) -> ApiResult:
    query = urllib.parse.urlencode(params)
    target = f"{url}?{query}"
    req = urllib.request.Request(target, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            raw = response.read().decode("utf-8")
            return ApiResult(True, response.status, json.loads(raw or "{}"))
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8")
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            data = {"raw": raw}
        return ApiResult(False, exc.code, data)


def insight_values(payload: dict[str, Any]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for item in payload.get("data", []):
        name = item.get("name")
        value_list = item.get("values") or []
        if not name:
            continue
        if value_list:
            values[name] = value_list[-1].get("value")
        else:
            values[name] = item.get("total_value", {}).get("value")
    return values


def collect_instagram_post(media_id: str, token: str) -> dict[str, Any]:
    metrics = ",".join(INSTAGRAM_POST_METRICS)
    result = request_json(
        f"{GRAPH_BASE}/{media_id}/insights",
        {"metric": metrics, "access_token": token},
    )
    if not result.ok:
        raise RuntimeError(
            json.dumps(
                {"platform": "instagram", "status": result.status, "response": result.data},
                ensure_ascii=False,
            )
        )
    return insight_values(result.data)


def collect_instagram_media_metadata(media_id: str, token: str) -> dict[str, Any]:
    result = request_json(
        f"{GRAPH_BASE}/{media_id}",
        {
            "fields": "id,caption,media_type,permalink,timestamp,like_count,comments_count",
            "access_token": token,
        },
    )
    if not result.ok:
        raise RuntimeError(
            json.dumps(
                {"platform": "instagram", "status": result.status, "response": result.data},
                ensure_ascii=False,
            )
        )
    return result.data


def collect_instagram_account_snapshot(ig_user_id: str, token: str) -> dict[str, Any]:
    result = request_json(
        f"{GRAPH_BASE}/{ig_user_id}",
        {"fields": "id,username,followers_count,media_count", "access_token": token},
    )
    if not result.ok:
        raise RuntimeError(
            json.dumps(
                {"platform": "instagram_account", "status": result.status, "response": result.data},
                ensure_ascii=False,
            )
        )
    return result.data


def collect_threads_post(thread_id: str, token: str) -> dict[str, Any]:
    metrics = ",".join(THREADS_POST_METRICS)
    result = request_json(
        f"{THREADS_BASE}/{thread_id}/insights",
        {"metric": metrics, "access_token": token},
    )
    if not result.ok:
        raise RuntimeError(
            json.dumps(
                {"platform": "threads", "status": result.status, "response": result.data},
                ensure_ascii=False,
            )
        )
    return insight_values(result.data)


def collect_threads_metadata(thread_id: str, token: str) -> dict[str, Any]:
    result = request_json(
        f"{THREADS_BASE}/{thread_id}",
        {"fields": "id,permalink,text,timestamp,media_type", "access_token": token},
    )
    if not result.ok:
        raise RuntimeError(
            json.dumps(
                {"platform": "threads_metadata", "status": result.status, "response": result.data},
                ensure_ascii=False,
            )
        )
    return result.data


def append_row(path: Path, sheet_name: str, values: list[Any]) -> None:
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    worksheet.append(values)
    workbook.save(path)


def append_instagram(path: Path, media_id: str, campaign: str, content_type: str) -> None:
    token = os.environ.get("INSTAGRAM_ACCESS_TOKEN")
    ig_user_id = os.environ.get("INSTAGRAM_IG_USER_ID")
    if not token or not ig_user_id:
        raise SystemExit("Missing INSTAGRAM_ACCESS_TOKEN or INSTAGRAM_IG_USER_ID")

    metadata = collect_instagram_media_metadata(media_id, token)
    insights = collect_instagram_post(media_id, token)
    account = collect_instagram_account_snapshot(ig_user_id, token)
    now = datetime.now(timezone.utc).isoformat()

    followers_gained_note = "Derive from account followers_count delta before/after publish window."
    append_row(
        path,
        "instagram_posts",
        [
            now[:10],
            media_id,
            metadata.get("permalink", ""),
            metadata.get("timestamp", ""),
            campaign,
            content_type,
            metadata.get("caption", ""),
            insights.get("reach", ""),
            metadata.get("like_count", insights.get("likes", "")),
            metadata.get("comments_count", insights.get("comments", "")),
            insights.get("saved", ""),
            insights.get("shares", ""),
            "",
            account.get("followers_count", ""),
            followers_gained_note,
            now,
        ],
    )


def append_threads(path: Path, thread_id: str, campaign: str, content_type: str) -> None:
    token = os.environ.get("THREADS_ACCESS_TOKEN")
    if not token:
        raise SystemExit("Missing THREADS_ACCESS_TOKEN")

    metadata = collect_threads_metadata(thread_id, token)
    insights = collect_threads_post(thread_id, token)
    now = datetime.now(timezone.utc).isoformat()
    append_row(
        path,
        "threads_posts",
        [
            now[:10],
            thread_id,
            metadata.get("permalink", ""),
            metadata.get("timestamp", ""),
            campaign,
            content_type,
            metadata.get("text", ""),
            insights.get("views", ""),
            insights.get("likes", ""),
            insights.get("replies", ""),
            insights.get("reposts", ""),
            insights.get("quotes", ""),
            insights.get("shares", ""),
            now,
        ],
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Collect social metrics into social_metrics.xlsx")
    parser.add_argument("--workbook", default=os.environ.get("SOCIAL_METRICS_XLSX", "social_metrics.xlsx"))
    parser.add_argument("--instagram-media-id")
    parser.add_argument("--threads-post-id")
    parser.add_argument("--campaign", default="")
    parser.add_argument("--content-type", default="")
    args = parser.parse_args()

    path = Path(args.workbook).expanduser().resolve()
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    if args.instagram_media_id:
        append_instagram(path, args.instagram_media_id, args.campaign, args.content_type)

    if args.threads_post_id:
        time.sleep(1)
        append_threads(path, args.threads_post_id, args.campaign, args.content_type)

    if not args.instagram_media_id and not args.threads_post_id:
        print("No collection target provided. Nothing executed.")
    else:
        print(f"Metrics appended to {path}")

    print(
        json.dumps(
            {
                "instagram_token": mask(os.environ.get("INSTAGRAM_ACCESS_TOKEN")),
                "threads_token": mask(os.environ.get("THREADS_ACCESS_TOKEN")),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
